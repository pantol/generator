#!/usr/bin/env python3

import csv
import random
import datetime
from openpyxl import load_workbook

def generate_random_date(start_date, end_date):
    return start_date + datetime.timedelta(days=random.randint(0, (end_date - start_date).days))

def generate_invoices(csv_file, nips, invoices_per_nip, prefix_inv, start_date, end_date, start_date_wymg, end_date_wymg, waluta_faktur):
    used_invoice_numbers = set()
    with open(csv_file, mode='w', newline='') as file:
        writer = csv.writer(file, delimiter=';')
        writer.writerow(['#SOF', str(datetime.date.today()), '1'])
        
        for i, nip in enumerate(nips, 1):
            for _ in range(invoices_per_nip):
                invoice_number = random.randint(1, len(nips) * invoices_per_nip)
                while invoice_number in used_invoice_numbers:
                    invoice_number = random.randint(1, len(nips) * invoices_per_nip)
                used_invoice_numbers.add(invoice_number)
                
                writer.writerow([
                    nip,
                    f"{prefix_inv}{invoice_number}",
                    generate_random_date(start_date, end_date),
                    generate_random_date(start_date_wymg, end_date_wymg),
                    waluta_faktur,
                    1000,  # Brutto
                    0,     # VAT
                    1000   # saldo_faktury
                ])
            
            if i % 1000 == 0:
                print(f"Generated {i * invoices_per_nip} invoices")
        
        writer.writerow(['#EOF', str(len(nips) * invoices_per_nip)])

def get_nips():
    choice = input("Do you want to (1) read NIPs from a file or (2) use a hardcoded list? Enter 1 or 2: ")
    if choice == "1":
        file_path = input("Enter the path to the Excel file containing NIPs: ")
        wb = load_workbook(filename=file_path, read_only=True)
        return [str(row[0]) for row in wb.active.iter_rows(values_only=True) if row[0]]
    else:
        return [
            # ... add 10,000 hardcoded NIPs here ...
        ]

def main():
    nips = get_nips()
    if len(nips) != 10000:
        print(f"Warning: The number of NIPs provided ({len(nips)}) is not 10,000.")
        if input("Do you want to proceed? (y/n): ").lower() != 'y':
            return

    generate_invoices(
        "data_10000_nips2.csv",
        nips,
        7,  # invoices_per_nip
        input('Podaj prefix dla faktur w pliku: ') + '/',
        datetime.datetime.strptime(input('Zakres początkowy dat wystawienia faktur (YYYY-MM-DD): '), '%Y-%m-%d').date(),
        datetime.datetime.strptime(input('Zakres końcowy dat wystawienia faktur (YYYY-MM-DD): '), '%Y-%m-%d').date(),
        datetime.datetime.strptime(input('Zakres początkowy dat wymagalności faktur (YYYY-MM-DD): '), '%Y-%m-%d').date(),
        datetime.datetime.strptime(input('Zakres końcowy dat wymagalności faktur (YYYY-MM-DD): '), '%Y-%m-%d').date(),
        input('Wpisz walutę faktury: ').upper()
    )

    print(f"Plik CSV 'data_10000_nips.csv' wygenerowany z {len(nips) * 7} fakturami dla {len(nips)} unikalnych NIP-ów.")

if __name__ == "__main__":
    main()